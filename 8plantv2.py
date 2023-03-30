# Import libraries 
from plantcv import plantcv as pcv

from tkinter import Tk
from tkinter.filedialog import askopenfilenames

from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl 

import pathlib

import numpy as np
from os.path import exists

from os import chdir
current_file = pathlib.Path(__file__).parent.resolve()
chdir (current_file)#change thsi to change the drive or folder you are working in

#-----------------------------------------------------------------------------
#This is the setup for the program like asking he name and asking for images input
plant_id = input("Name of plant? >")

Tk().withdraw()
plant_imgs = askopenfilenames()

ammount_of_imgs = len(plant_imgs)

print(f"Ammount of imgs> {len(plant_imgs)}")

for i in plant_imgs:
    print (i)

black_data = []
imgnum = 0


class options():
    def __init__(self):
        self.image = plant_imgs[imgnum]
        self.debug = "plot"
        self.writeimg= False
        self.outdir = "*"
print("step 1 of 4 done")

#------------------------------------------------------------------------------
#this makes the black and white images form the inputed plant images
for i in range (ammount_of_imgs): #change this number in order to change how many plants are scaned
    # Get options
    args = options()
    
        # Set debug to the global parameter 
    #pcv.params.debug = args.debug
        
        
        # Inputs:
        #   filename - Image file to be read in 
        #   mode - How to read in the image; either 'native' (default), 'rgb', 'gray', or 'csv'
    img, path, filename = pcv.readimage(filename=args.image)
        
        # Inputs:
        #   rgb_img = original color image
        #   channel = one of the colorspaces: 'l', 'a', or 'b'
    a = pcv.rgb2gray_lab(rgb_img = img, channel='a')
    
        # Inputs:
        #   gray_img    = Grayscale image data
        #   threshold   = Threshold value (0-255)
        #   max_value   = Value to apply above threshold (255 = white)
        #   object_type = "light" or "dark" (default: "light"). If object is lighter than the background then standard 
        #                 thresholding is done. If object is darker than t
    
    black_data.append(pcv.threshold.binary(gray_img=a, threshold=119, max_value=255, object_type='dark') )
    
    imgnum += 1

print ("step 2 of 4 done")   
    
#-----------------------------------------------------------------------------
#This dose the mathmatical operations on the data
def Average(lst):
    return sum(lst) / len(lst)
white_pix = []
black_pix = []
percent_pix = []
all_pix = 48000000
for i in range(ammount_of_imgs): 
    white_pix.append (np.sum(black_data[i] == 255))
    percent_pix.append (float(100) * (white_pix[i]/all_pix))
    average_percent = Average(percent_pix)
    average_white = Average(white_pix)
    
print ("step 3 of 4 done")
    
    
#------------------------------------------------------------------------------------
#This writes the data that has been gathered to the excel document
if exists(str(current_file) + "\\" + "img_data.xlsx") == False: #change this to change the drive or folder you are working in
    workbook = Workbook()
    workbook.save("img_data.xlsx")# change this to chage the output file
    workbook.close()
        

wb = openpyxl.load_workbook("img_data.xlsx") # change this to chage the output file
  
sheet = wb.active 

data = (
    ["--", "--", "--", "--", "--"],
    [plant_id],
    ('Filename','Percent', 'Total'),
    ('Average', average_percent, average_white)
    )


for row in data:
    sheet.append(row)

wb.save('img_data.xlsx')


workbook_name = 'img_data.xlsx'
wb = load_workbook(workbook_name)
page = wb.active







for n in range(ammount_of_imgs):
    Data2 = (
        [plant_imgs[n], percent_pix[n], white_pix[n]]
        )
    page.append(Data2)

wb.save('img_data.xlsx')





wb.close()
    
print ("step 4 of 4 done")






